from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

signals = [
    "END",
    "HLT",
    "PC_EN",
    "MDR_EN",
    "A_EN",
    "B_EN",
    "C_EN",
    "TMP_EN",
    "ALU_EN",
    "SP_EN",
    "IO_EN",
    "PC_LOAD",
    "PC_INR",
    "IR_LOAD",
    "A_LOAD",
    "A_INR",
    "A_DCR",
    "B_LOAD",
    "B_INR",
    "B_DCR",
    "C_LOAD",
    "C_INR",
    "C_DCR",
    "TMP_LOAD",
    "TMP_INR",
    "TMP_DCR",
    "OUT_LOAD",
    "SP_LOAD",
    "SP_INR",
    "SP_DCR",
    "MAR_LOAD",
    "MDR_LOADB",
    "MDR_LOADL",
    "MDR_LOADH",
    "RAM_WRITE",
    "ALU_OP3",
    "ALU_OP2",
    "ALU_OP1",
    "ALU_OP0",
    "FLAG_LOAD",
    "IO_LOAD",
    "UNUSED5",
    "UNUSED4",
    "UNUSED3",
    "UNUSED2",
    "UNUSED1",
    "UNUSED0",
]

instruction_info = {
    "80": {
        "Instruction": "ADD B",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "81": {
        "Instruction": "ADD C",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "A0": {
        "Instruction": "ANA B",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "A1": {
        "Instruction": "ANA C",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "E6": {
        "Instruction": "ANI byte",
        "Flags": "S,Z",
        "Addressing": "Immediate",
        "Bytes": 2,
    },
    "CD": {
        "Instruction": "CALL addr",
        "Flags": "None",
        "Addressing": "Immediate",
        "Bytes": 3,
    },
    "2F": {"Instruction": "CMA", "Flags": "None", "Addressing": "Implied", "Bytes": 1},
    "3D": {
        "Instruction": "DCR A",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "05": {
        "Instruction": "DCR B",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "0D": {
        "Instruction": "DCR C",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "76": {"Instruction": "HLT", "Flags": "None", "Addressing": "-", "Bytes": 1},
    "DB": {
        "Instruction": "IN byte",
        "Flags": "None",
        "Addressing": "Direct",
        "Bytes": 2,
    },
    "3C": {
        "Instruction": "INR A",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "04": {
        "Instruction": "INR B",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "0C": {
        "Instruction": "INR C",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "FA": {
        "Instruction": "JM addr",
        "Flags": "None",
        "Addressing": "Immediate",
        "Bytes": 3,
    },
    "C3": {
        "Instruction": "JMP addr",
        "Flags": "None",
        "Addressing": "Immediate",
        "Bytes": 3,
    },
    "C2": {
        "Instruction": "JNZ addr",
        "Flags": "None",
        "Addressing": "Immediate",
        "Bytes": 3,
    },
    "CA": {
        "Instruction": "JZ addr",
        "Flags": "None",
        "Addressing": "Immediate",
        "Bytes": 3,
    },
    "3A": {
        "Instruction": "LDA addr",
        "Flags": "None",
        "Addressing": "Direct",
        "Bytes": 3,
    },
    "78": {
        "Instruction": "MOV A,B",
        "Flags": "None",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "79": {
        "Instruction": "MOV A,C",
        "Flags": "None",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "47": {
        "Instruction": "MOV B,A",
        "Flags": "None",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "41": {
        "Instruction": "MOV B,C",
        "Flags": "None",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "4F": {
        "Instruction": "MOV C,A",
        "Flags": "None",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "48": {
        "Instruction": "MOV C,B",
        "Flags": "None",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "3E": {
        "Instruction": "MVI A,byte",
        "Flags": "None",
        "Addressing": "Immediate",
        "Bytes": 2,
    },
    "06": {
        "Instruction": "MVI B,byte",
        "Flags": "None",
        "Addressing": "Immediate",
        "Bytes": 2,
    },
    "0E": {
        "Instruction": "MVI C,byte",
        "Flags": "None",
        "Addressing": "Immediate",
        "Bytes": 2,
    },
    "00": {"Instruction": "NOP", "Flags": "None", "Addressing": "-", "Bytes": 1},
    "B0": {
        "Instruction": "ORA B",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "B1": {
        "Instruction": "ORA C",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "F6": {
        "Instruction": "ORI byte",
        "Flags": "S,Z",
        "Addressing": "Immediate",
        "Bytes": 2,
    },
    "D3": {
        "Instruction": "OUT byte",
        "Flags": "None",
        "Addressing": "Direct",
        "Bytes": 2,
    },
    "17": {"Instruction": "RAL", "Flags": "None", "Addressing": "Implied", "Bytes": 1},
    "1F": {"Instruction": "RAR", "Flags": "None", "Addressing": "Implied", "Bytes": 1},
    "C9": {"Instruction": "RET", "Flags": "None", "Addressing": "Implied", "Bytes": 1},
    "32": {
        "Instruction": "STA addr",
        "Flags": "None",
        "Addressing": "Direct",
        "Bytes": 3,
    },
    "90": {
        "Instruction": "SUB B",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "91": {
        "Instruction": "SUB C",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "A8": {
        "Instruction": "XRA B",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "A9": {
        "Instruction": "XRA C",
        "Flags": "S,Z",
        "Addressing": "Register",
        "Bytes": 1,
    },
    "EE": {
        "Instruction": "XRI byte",
        "Flags": "S,Z",
        "Addressing": "Immediate",
        "Bytes": 2,
    },
}

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

wb = Workbook()
ws = wb.active
ws.title = "Control Words"

ws.append(["Opcode", "Instruction", "Flags", "Addressing", "Bytes", "Stage"] + signals)

for opcode in range(256):
    hex_op = f"{opcode:02X}"
    if hex_op in instruction_info:
        info = instruction_info[hex_op]

        for stage in range(16):
            row = [
                hex_op,
                info["Instruction"],
                info["Flags"],
                info["Addressing"],
                info["Bytes"],
                stage,
            ]

            # default all signals off, END=1
            signal_values = ["0"] * len(signals)
            signal_values[signals.index("END")] = "1"

            if stage == 0:  # T1: PC -> MAR
                signal_values[signals.index("END")] = "0"
                signal_values[signals.index("PC_EN")] = "1"
                signal_values[signals.index("MAR_LOAD")] = "1"

            elif stage == 1:  # T2: RAM[MAR] -> MDR
                signal_values[signals.index("END")] = "0"
                signal_values[signals.index("MDR_LOADL")] = "1"

            elif stage == 2:  # T3: MDR -> IR, PC <- PC+1
                signal_values[signals.index("END")] = "0"
                signal_values[signals.index("MDR_EN")] = "1"
                signal_values[signals.index("IR_LOAD")] = "1"
                signal_values[signals.index("PC_INR")] = "1"

            ws.append(row + signal_values)

            # Highlight all "1" cells in green
            for col_idx, val in enumerate(signal_values, start=len(row) + 1):
                if val == "1":
                    ws.cell(ws.max_row, col_idx).fill = green_fill

    else:
        for stage in range(16):
            row = [hex_op, "", "", "", "", stage]
            signal_values = ["0"] * len(signals)
            signal_values[signals.index("END")] = "0"
            ws.append(row + signal_values)

            # Highlight all "1" cells in green
            for col_idx, val in enumerate(signal_values, start=len(row) + 1):
                if val == "1":
                    ws.cell(ws.max_row, col_idx).fill = green_fill

# Add conditional formatting rule for signal value columns (columns G onwards)
signal_start_col = 7  # Column G (1-indexed: A=1, B=2, ..., G=7)
signal_end_col = signal_start_col + len(signals) - 1

# Create conditional formatting rule: if cell value equals "1", fill green
rule = CellIsRule(operator="equal", formula=["1"], fill=green_fill)

# Apply the rule to all signal value columns
signal_range = f"{ws.cell(2, signal_start_col).coordinate}:{ws.cell(ws.max_row, signal_end_col).coordinate}"
ws.conditional_formatting.add(signal_range, rule)

wb.save("control_words_blank.xlsx")
print(
    "Generated control_words_blank.xlsx with green highlights and conditional formatting."
)
